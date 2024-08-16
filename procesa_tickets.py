import pandas as pd
import datetime
import warnings

# METODO DE CARGA PARA DESAROLLO

# valor_boleto = 45
# archivo_1 = r"C:\Users\bchavat\Desktop\dev_environments\automat_ticket_laboral\Solicitud de Tickets Alimentación _ Transporte_2.xlsx"
# archivo_2 = r"C:\Users\bchavat\Desktop\dev_environments\automat_ticket_laboral\Query.xlsx"
# archivo_salida = r"C:\Users\bchavat\Desktop\dev_environments\automat_ticket_laboral\procesa_ticket2407.xlsx"

ajustes = {
        'Número de Funcionario': [4113, 5585, 4239, 5036, 4225, 5741, 4351, 4112, 3974, 4488, 4794, 4993, 4814, 5346,
                                  4703, 6485, 2327, 6251, 5247, 5243, 3471, 3521, 2707, 5740, 1003, 5497, 3559, 1432,
                                  5023, 5086, 5896, 5805, 5709, 5668, 5665, 5624, 5017, 5267, 5213, 5140, 5475, 4792,
                                  4684, 4667, 4656, 4349, 4212, 3363, 3280, 2678, 1830, 1393, 1191, 4824, 3770, 4933,
                                  4129, 3989, 5224, 5239, 6575, 2397, 3916, 4446],
        'Documento identidad': [27543104, 46031302, 25331806, 50049252, 52455522, 46594039, 41855797, 35914761, 37529144,
                                36942785, 45022584, 46672059, 30746559, 41443340, 47514406, 34443599, 50658510, 39888425,
                                52663139, 53084445, 47171363, 34929622, 51297226, 40267159, 41833511, 46994269, 45247788,
                                45106348, 45215822, 44472633, 40227812, 39611660, 25843089, 39527104, 37929932, 46276643,
                                49057468, 25707655, 32959851, 38152239, 36048292, 20212277, 19462750, 43601233, 48968058,
                                51423803, 45333678, 36552075, 45891214, 42093138, 19854424, 13562320, 43261625, 49553901,
                                46748448, 0, 47637008, 18288200, 0, 0, 43748382, 49225910, 49254307, 40866276],
        'Apellido y Nombre': ['NUÑEZ MIRTA', 'RODRIGUEZ DIEGO', 'NUÑEZ DARDO', 'FREITAS LORENA', 'HORNOS DANIELA',
                              'ROLDAN FABRICIO', 'CABRERA OSCAR', 'ACOSTA ANA', 'MOREIRA MAICOL', 'SANDRA SILVA',
                              'LANDRIEU NATHALIE', 'BURGUEÑO LUCIA', 'GONZALEZ HUGO', 'PERI ADRIANA', 'GERMAN DIEGO',
                              'DE ABREU ALEXANDRA', 'GONZALEZ XIMENA', 'SANTOS OMAR', 'GONZALEZ AGUSTIN', 'GONZALEZ CRISTHIAN',
                              'ALVAREZ SANTIAGO', 'DA CUNDA ANA', 'CALABUIG MARIA', 'GONZALEZ PATRICIA', 'MOREIRA ALEXIS',
                              'MARIN ANDREA', 'ELIZALDE NICOLAS', 'SAMPAYO CRISTINA', 'CAPOTE LETICIA', 'GUTIERREZ ROMINA',
                              'RODRIGUEZ RUTH', 'SILVERA ADRIANA', 'GONZALEZ EDUARDO', 'DEL PERERA DIEGO', 'ARAUJO ALEJANDRO',
                              'LOPEZ JONATHAN', 'CASTRO JESSICA', 'ACUÑA MARIANELA', 'SANES MAURICIO', 'CORREA WASHINGTON',
                              'MALGOR MARIANA', 'ALVAREZ NICOLAS', 'BARRETO NATALI', 'DIAZ MARTIN', 'MORALES STEFANI',
                              'FIERRO VERONICA', 'DUARTE CRISTIAN', 'SANTOS SEBASTIAN', 'GOMEZ LUCIA', 'RUEDA DIEGO',
                              'CHICHIZIOLA JESUS', 'PISANI MARTA', 'CONDE RONEY', 'RIVERA DIEGO', 'GUGLIELMELLI DAHIANA',
                              'NALERIO FLORENCIA', 'SILVA ALBA', 'FERRER BEATRIZ', 'DE LA PUENTE RODRIGO', 'BAZZANO MARCELO',
                              'PAZ ALBERTINA', 'NUÑEZ NATALIA', 'ALVAREZ TORENA NATALIA LORELEY', 'BATISTA SANTO JUAN GONZALO'],
        'Ajuste julio 2016': [0.10, 0.12, 0.10, 0.12, 0.12, 0.12, 0.12, 0.10, 0.10, 0.12, 0.12, 0.12, 0.10, 0.10, 0.12,
                              0.00, 0.10, 0.10, 0.12, 0.12, 0.10, 0.10, 0.10, 0.12, 0.10, 0.10, 0.10, 0.10, 0.12, 0.12,
                              0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.12, 0.10, 0.12, 0.12, 0.10, 0.10, 0.12, 0.12, 0.12,
                              0.12, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
                              0.00, 0.00, 0.00, 0.00],
        'Ajuste enero 2017': [0.04, 0.06, 0.04, 0.06, 0.06, 0.06, 0.06, 0.04, 0.04, 0.06, 0.06, 0.06, 0.04, 0.04, 0.06,
                              0.04, 0.04, 0.04, 0.06, 0.06, 0.04, 0.04, 0.04, 0.06, 0.04, 0.04, 0.04, 0.04, 0.06, 0.06,
                              0.04, 0.04, 0.04, 0.04, 0.04, 0.04, 0.06, 0.04, 0.06, 0.06, 0.04, 0.04, 0.06, 0.06, 0.06,
                              0.06, 0.04, 0.04, 0.04, 0.04, 0.04, 0.04, 0.04, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
                              0.00, 0.00, 0.00, 0.00],
        'Ajuste julio 2017': [0.03, 0.06, 0.03, 0.06, 0.06, 0.06, 0.06, 0.03, 0.03, 0.06, 0.06, 0.06, 0.03, 0.03, 0.00,
                              0.03, 0.03, 0.03, 0.06, 0.06, 0.03, 0.03, 0.03, 0.06, 0.03, 0.03, 0.03, 0.03, 0.06, 0.06,
                              0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.06, 0.03, 0.06, 0.06, 0.03, 0.03, 0.06, 0.06, 0.06,
                              0.06, 0.03, 0.00, 0.03, 0.03, 0.03, 0.03, 0.03, 0.06, 0.03, 0.06, 0.06, 0.03, 0.06, 0.06,
                              0.03, 0.03, 0.06, 0.06]
    }
ajustes = pd.DataFrame(ajustes)
warnings.filterwarnings('ignore')
# -----------------------------------------------------------------------------------
def fun_procesa_tickets(carga_1, carga_2, guarda, vboleto):

    archivo_1 = pd.read_excel(carga_1)
    archivo_2 = pd.read_excel(carga_2)
    archivo_salida = guarda
    valor_boleto = int(vboleto)

    nomina = archivo_2.copy()

    nomina = nomina.loc[:,['Nº funcionario',
                    'Documento identidad',
                    'Apellidos y nombres',
                    'Centro costo (car)',
                    'Descripción cargo (car)'
                           ]
                    ]

    control = archivo_1.copy()
    control = pd.merge(control, nomina, left_on='Número de colaborador:', right_on='Nº funcionario', how='left')

    control = pd.merge(control, ajustes, left_on='Número de colaborador:',
                       right_on='Número de Funcionario', how='left')

    control["control"] = control['Número de colaborador:'] - control['Nº funcionario']
    control = control.rename(columns={"En caso solicitar boletos que no son urbanos (Montevideo), ingrese el valor:": "valor_dif"})

    contador = 0
    for valor_control in control["control"]:
        indice_control = control.index[contador]
        if valor_control != 0:
            control['control'][indice_control] = "Nro. colab. equivocado o no está en nomina."
        else:
            control['control'][indice_control] = "OK"
        contador = contador + 1

    control["Ajuste julio 2016"] = control["Ajuste julio 2016"].fillna(0)
    control["Ajuste enero 2017"] = control["Ajuste enero 2017"].fillna(0)
    control["Ajuste julio 2017"] = control["Ajuste julio 2017"].fillna(0)

    fecha_hoy = datetime.date.today().strftime('%d-%m-%Y')

    control[f"tickets_{fecha_hoy}"] = round(control["Cantidad:"] * (1 + control["Ajuste julio 2016"]) * (1 + control["Ajuste enero 2017"]) * (1 + control["Ajuste julio 2017"]))

    control["Costo"] = int(valor_boleto)
    control["valor_dif"] = control["valor_dif"].fillna(0).astype(int)
    for index in range(0, len(control["Costo"])):
        if control.loc[index, "valor_dif"] != 0:
            control.loc[index, "Costo"] = control.loc[index, "valor_dif"]

    control["Monto"] = control[f"tickets_{fecha_hoy}"] * control["Costo"]

    final = control[['Usuario',
                     'Fecha',
                     'Sucursal:',
                     'Número de colaborador:',
                     'Apellido y Nombre (escribir primero el apellido):',
                     'Nº funcionario',
                     'Descripción cargo (car)',
                     'Apellidos y nombres',
                     'Tipo de tickets:',
                     'Cantidad:',
                     'Ajuste julio 2016',
                     'Ajuste enero 2017',
                     'Ajuste julio 2017',
                     f'tickets_{fecha_hoy}',
                     'Costo',
                     'Monto',
                     'Motivo de solicitud:',
                     'Observaciones:',
                     'control']]

    nombres_columnas = {'Apellidos y nombres': 'SAICO Apellidos y Nombres',
                        'Apellido y Nombre (escribir primero el apellido):': 'Apellido y Nombre',
                        'Nº funcionario': 'SAICO Nro Colab',
                        'Descripción cargo (car)': 'SAICO Descripcion cargo',

                        'control': 'Obvservacion automat'}

    final = final.rename(columns=nombres_columnas)

    # final.to_excel(archivo_salida, index=False)

    #-----------------FORMATEO DE DATAFRAME---------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Crear libro de trabajo de Excel
    workbook = Workbook()
    hoja_1 = workbook.active

    for fila in dataframe_to_rows(final, index=False, header=True):
        hoja_1.append(fila)

    # print(hoja_1.max_row, hoja_1.max_column)

    def conv_num_a_letra(numero):
        letras = ""
        while numero > 0:
            numero -= 1
            letras = chr((numero % 26) + 65) + letras
            numero //= 26
        return letras

    def letra_a_numero(columna):
        resultado = 0
        for i, letra in enumerate(reversed(columna)):
            resultado += (ord(letra.upper()) - ord('A') + 1) * (26 ** i)
        return resultado

    def selecciona_celdas(rango=None,
                          sheet=None,
                          value=None,
                          font=None,
                          alignment=None,
                          border=None,
                          fill=None,
                          number_format=None,
                          hyperlink=None,
                          comment=None,
                          rule=None):

        r = rango.split(":")
        row = []
        columns = []

        for i in range(0, 2):
            col_num = []
            row_num = []
            for caracter in r[i]:
                if caracter.isalpha():
                    letras = caracter
                    num = letra_a_numero(letras)
                    col_num.append(str(num))

                elif caracter.isdigit():
                    row_num.append(str(caracter))

            columns.append(''.join(col_num))
            row.append(''.join(row_num))

        col_start = int(columns[0])
        col_end = int(columns[1]) + 1
        row_start = int(row[0])
        row_end = int(row[1]) + 1

        for col in range(col_start, col_end):
            for row in range(row_start, row_end):
                cell = sheet.cell(row=row, column=col)
                if value is not None:
                    cell.value = value

                if font is not None:
                    cell.font = font

                if alignment is not None:
                    cell.alignment = alignment

                if border is not None:
                    cell.border = border

                if fill is not None:
                    cell.fill = fill

                if number_format is not None:
                    cell.number_format = number_format

                if hyperlink is not None:
                    cell.hyperlink = hyperlink

                if comment is not None:
                    cell.comment = comment

                if rule is not None:
                    cell.rule = rule

    max_col = hoja_1.max_column
    max_col = conv_num_a_letra(max_col)

    borde_param = Side(style='thin', color="808080")
    selecciona_celdas(rango=f"A1:{max_col}{hoja_1.max_row}",
                      sheet=hoja_1,
                      font=Font(name='Calibri Light'),
                      alignment=Alignment(horizontal='center', vertical='center'),
                      border=Border(left=borde_param, right=borde_param,
                                    top=borde_param, bottom=borde_param))

    color_header = "204E60"
    color_title = "FFFFFF"
    selecciona_celdas(rango=f"A1:{max_col}1",
                      sheet=hoja_1,
                      font=Font(bold=True, color=color_title),
                      fill=PatternFill(start_color=color_header, fill_type="solid"))

    for column in hoja_1.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

            adjusted_width = (max_length + 4)
            hoja_1.column_dimensions[cell.column_letter].width = adjusted_width

    workbook.save(archivo_salida + ".xlsx")

# fun_procesa_tickets(archivo_1, archivo_2, archivo_salida, 45)






